"""
backtest_report_exporter.py — write backtest results to Excel.

One sheet of per-day rows plus an aggregate summary block. Output:
<repo_root>/runtime/reports/backtest_<start>_<n>d.xlsx
"""
from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.styles import Font

from phase_6_backtesting_and_validation.backtest_engine.backtest_runner import BacktestResult


def _repo_root() -> str:
    # this file: <repo>/phase_6_backtesting_and_validation/backtest_excel_reports/ -> up 2
    here = os.path.dirname(os.path.abspath(__file__))
    return os.path.abspath(os.path.join(here, os.pardir, os.pardir))


def export_backtest(start_date: str, result: BacktestResult) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Backtest"
    headers = ["date", "feasible", "checker_pass", "objective_eur", "solve_sec",
               "price_mae", "price_rmse", "pv_mae", "note"]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
    for r in result.rows:
        ws.append([r[h] for h in headers])

    ws2 = wb.create_sheet("Summary")
    bold = Font(bold=True)
    ws2["A1"] = f"Backtest summary — {start_date}, {result.n_days} days"
    ws2["A1"].font = Font(bold=True, size=13)
    for i, (label, val) in enumerate([
        ("Days", result.n_days),
        ("Feasible", result.n_feasible),
        ("Checker passed", result.n_checker_pass),
        ("Avg objective (EUR)", round(result.avg_objective_eur, 2)),
        ("Avg solve (s)", round(result.avg_solve_sec, 3)),
        ("Avg price MAE (EUR/MWh)", round(result.avg_price_mae, 2)),
        ("Avg PV MAE (MW)", round(result.avg_pv_mae, 4)),
    ], start=3):
        ws2[f"A{i}"] = label; ws2[f"A{i}"].font = bold
        ws2[f"B{i}"] = val

    # --- Risk sheet ---
    if result.risk is not None:
        wr = wb.create_sheet("Risk")
        wr["A1"] = "Portfolio Risk Metrics"
        wr["A1"].font = Font(bold=True, size=13)
        wr["A2"] = f"Source: {result.risk.n_days} feasible backtest days  |  " \
                   f"Bootstrap n=10,000  |  alpha=95% and 99%"

        risk_rows = [
            ("", ""),
            ("--- P&L Distribution ---", ""),
            ("Mean daily P&L (EUR)",              result.risk.mean_pnl_eur),
            ("Std daily P&L (EUR)",               result.risk.std_pnl_eur),
            ("Min daily P&L (EUR)",               result.risk.min_pnl_eur),
            ("Max daily P&L (EUR)",               result.risk.max_pnl_eur),
            ("", ""),
            ("--- Historical Simulation ---", ""),
            ("VaR(95%)  — 5th-pct P&L (EUR)",    result.risk.var_95_eur),
            ("CVaR(95%) — Expected Shortfall (EUR)", result.risk.cvar_95_eur),
            ("VaR(99%)  — 1st-pct P&L (EUR)",    result.risk.var_99_eur),
            ("CVaR(99%) — Expected Shortfall (EUR)", result.risk.cvar_99_eur),
            ("", ""),
            ("--- Monte Carlo Bootstrap (VaR 95%) ---", ""),
            ("VaR(95%)  mean  (EUR)",             result.risk.var_95_mean),
            ("VaR(95%)  std   (EUR)  ± CI",       result.risk.var_95_std),
            ("CVaR(95%) mean  (EUR)",             result.risk.cvar_95_mean),
            ("CVaR(95%) std   (EUR)  ± CI",       result.risk.cvar_95_std),
            ("", ""),
            ("--- Risk-Adjusted ---", ""),
            ("Sharpe ratio (annualised, rf=0)",   result.risk.sharpe_ratio),
            ("Max drawdown (EUR)",                result.risk.max_drawdown_eur),
        ]
        for i, (label, val) in enumerate(risk_rows, start=4):
            wr[f"A{i}"] = label
            if label.startswith("---"):
                wr[f"A{i}"].font = Font(bold=True)
            if val != "":
                wr[f"B{i}"] = val

    # --- Operational analytics sheet (per-day summary) ---
    feasible_rows = [r for r in result.rows if r.get("feasible") and r.get("ops")]
    if feasible_rows:
        wo = wb.create_sheet("Operational")
        ops_headers = [
            "date", "turbine_hours_total", "pump_hours_total",
            "turbine_starts_total", "pump_starts_total",
            "turb_avg_run_h", "turb_max_run_h",
            "pump_avg_run_h", "pump_max_run_h",
            "turb_hours_top25pct_price", "pump_hours_bot25pct_price",
            "bess_charge_hours", "bess_discharge_hours",
            "avg_units_turbining", "avg_units_pumping",
        ]
        wo.append(ops_headers)
        for c in wo[1]: c.font = Font(bold=True)
        for r in feasible_rows:
            ops = r.get("ops", {})
            wo.append([r["date"]] + [ops.get(k, "") for k in ops_headers[1:]])

        # Temporal split sheet
        wt = wb.create_sheet("Temporal")
        tmp_headers = ["date", "band",
                       "hours", "turbine_pct", "pump_pct",
                       "avg_net_mw", "avg_profit_eur_h", "avg_price_eur_mwh"]
        wt.append(tmp_headers)
        for c in wt[1]: c.font = Font(bold=True)
        for r in feasible_rows:
            tmp = r.get("tmp", {})
            for band in ("night", "morning", "afternoon", "evening"):
                bd = tmp.get(band, {})
                if not bd:
                    continue
                wt.append([r["date"], band] + [bd.get(k, "") for k in tmp_headers[2:]])

        # Extended KPI sheet
        we = wb.create_sheet("KPI_Extended")
        eco_headers = [
            "date",
            "turbine_capacity_factor_pct", "pump_capacity_factor_pct",
            "bess_discharge_cf_pct", "pv_utilisation_pct",
            "avg_turbine_efficiency_pct", "avg_pump_efficiency_pct",
            "reservoir_fill_end_pct",
            "head_min_m", "head_max_m", "head_range_m",
            "da_revenue_share_pct", "frr_revenue_share_pct",
            "energy_revenue_eur",
        ]
        we.append(eco_headers)
        for c in we[1]: c.font = Font(bold=True)
        for r in feasible_rows:
            eco = r.get("eco", {})
            we.append([r["date"]] + [eco.get(k, "") for k in eco_headers[1:]])

    out_dir = os.path.join(_repo_root(), "runtime", "reports")
    os.makedirs(out_dir, exist_ok=True)
    path = os.path.join(out_dir, f"backtest_{start_date}_{result.n_days}d.xlsx")
    wb.save(path)
    return path
