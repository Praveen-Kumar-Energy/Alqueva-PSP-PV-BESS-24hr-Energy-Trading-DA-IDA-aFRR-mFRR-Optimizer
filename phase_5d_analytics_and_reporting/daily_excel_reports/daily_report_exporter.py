"""
daily_report_exporter.py — write the industrial daily trading report to Excel.

Five sheets:
  1. Dispatch_Hourly  — 24-row × ~93-col full component dispatch table
  2. ISP_Activation   — 96-ISP aFRR/mFRR activation breakdown
  3. Gate_Decisions   — one row per trading gate (DA/IDA1/IDA2/IDA3/XBID)
  4. Summary_KPIs     — 10 KPI sections with section headers
  5. Glossary         — column-name → description + unit lookup

Output: <repo_root>/runtime/reports/daily_report_<date>.xlsx
"""
from __future__ import annotations

import os
from typing import List, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from common_layer.database import (
    PositionStore, ReserveStore, ActivationStore, ComponentStore,
)
from phase_5d_analytics_and_reporting.daily_excel_reports.dispatch_sheet_builder import (
    build_dispatch_hourly,
)
from phase_5d_analytics_and_reporting.daily_excel_reports.summary_kpi_builder import (
    build_summary_kpis,
)


# ── styling constants ─────────────────────────────────────────────────────────
_TITLE_FONT     = Font(bold=True, size=14, color="FFFFFFFF")
_HEADER_FONT    = Font(bold=True, color="FFFFFFFF")
_SECTION_FONT   = Font(bold=True, color="FF1F3864")
_BOLD           = Font(bold=True)
_OK_FILL        = PatternFill("solid", fgColor="FFC6EFCE")
_BAD_FILL       = PatternFill("solid", fgColor="FFFFC7CE")
_HEADER_FILL    = PatternFill("solid", fgColor="FF1F3864")  # dark navy
_TITLE_FILL     = PatternFill("solid", fgColor="FF1F497D")  # slightly lighter
_ALT_FILL       = PatternFill("solid", fgColor="FFF2F2F2")  # light grey alt rows
_THIN           = Side(style="thin", color="FFBFBFBF")
_THIN_BORDER    = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER         = Alignment(horizontal="center", vertical="center", wrap_text=False)
_WRAP           = Alignment(wrap_text=True, vertical="top")

# Column group colours (for Dispatch_Hourly header band)
_GROUP_FILLS = {
    "A": "FFDAE8FC",   # blue-grey  — inputs
    "B": "FFD5E8D4",   # green      — PSP totals
    "C": "FFE1F5FE",   # light blue — per-unit PSP
    "D": "FFFFE0B2",   # orange     — PV
    "E": "FFFFF9C4",   # yellow     — BESS
    "F": "FFF3E5F5",   # lilac      — reservoir
    "G": "FFE8F5E9",   # mint       — efficiency
    "H": "FFFFF3E0",   # cream      — IDA
    "I": "FFFE9898",   # rose       — aFRR
    "J": "FFB9B9FF",   # lavender   — mFRR
    "K": "FFD0E0E3",   # teal       — headroom
    "L": "FFE2EFDA",   # pale green — revenue
}

# Maps column names → (group_letter, description, unit)
_COL_META = {
    "Hour":                    ("A", "Delivery hour (1–24)", "h"),
    "DA_price_EUR_MWh":        ("A", "OMIE DA forecast price", "EUR/MWh"),
    "PV_available_MW":         ("A", "PV irradiance forecast (available)", "MW"),
    "Reservoir_inflow_m3h":    ("A", "Natural inflow to upper reservoir", "m³/h"),
    "DA_side":                 ("B", "DA bid side: SELL/BUY/IDLE", "—"),
    "PSP_gen_MW":              ("B", "Total PSP turbine generation", "MW"),
    "PSP_pump_MW":             ("B", "Total PSP pump load", "MW"),
    "Plant_net_DA_MW":         ("B", "Total plant DA bid: PSP_gen−PSP_pump+PV_used+BESS_net. Positive=sell.", "MW"),
    "Plant_net_final_MW":      ("B", "Final committed position after all IDA/XBID adjustments", "MW"),
    "Units_turbining":         ("B", "Number of units turbining", "—"),
    "Units_pumping":           ("B", "Number of units pumping", "—"),
    "PSP_gen_u1_MW":           ("C", "Unit 1 turbine output", "MW"),
    "PSP_gen_u2_MW":           ("C", "Unit 2 turbine output", "MW"),
    "PSP_gen_u3_MW":           ("C", "Unit 3 turbine output", "MW"),
    "PSP_gen_u4_MW":           ("C", "Unit 4 turbine output", "MW"),
    "PSP_pump_u1_MW":          ("C", "Unit 1 pump input", "MW"),
    "PSP_pump_u2_MW":          ("C", "Unit 2 pump input", "MW"),
    "PSP_pump_u3_MW":          ("C", "Unit 3 pump input", "MW"),
    "PSP_pump_u4_MW":          ("C", "Unit 4 pump input", "MW"),
    "On_turb_u1":              ("C", "Unit 1 turbine on (binary)", "0/1"),
    "On_turb_u2":              ("C", "Unit 2 turbine on (binary)", "0/1"),
    "On_turb_u3":              ("C", "Unit 3 turbine on (binary)", "0/1"),
    "On_turb_u4":              ("C", "Unit 4 turbine on (binary)", "0/1"),
    "On_pump_u1":              ("C", "Unit 1 pump on (binary)", "0/1"),
    "On_pump_u2":              ("C", "Unit 2 pump on (binary)", "0/1"),
    "On_pump_u3":              ("C", "Unit 3 pump on (binary)", "0/1"),
    "On_pump_u4":              ("C", "Unit 4 pump on (binary)", "0/1"),
    "q_turb_u1_m3h":           ("C", "Unit 1 turbine water flow", "m³/h"),
    "q_turb_u2_m3h":           ("C", "Unit 2 turbine water flow", "m³/h"),
    "q_turb_u3_m3h":           ("C", "Unit 3 turbine water flow", "m³/h"),
    "q_turb_u4_m3h":           ("C", "Unit 4 turbine water flow", "m³/h"),
    "q_pump_u1_m3h":           ("C", "Unit 1 pump water flow", "m³/h"),
    "q_pump_u2_m3h":           ("C", "Unit 2 pump water flow", "m³/h"),
    "q_pump_u3_m3h":           ("C", "Unit 3 pump water flow", "m³/h"),
    "q_pump_u4_m3h":           ("C", "Unit 4 pump water flow", "m³/h"),
    "q_turb_total_m3h":        ("C", "Total turbine water flow (all units)", "m³/h"),
    "q_pump_total_m3h":        ("C", "Total pump water flow (all units)", "m³/h"),
    "PV_used_MW":              ("D", "PV directly fed to DA position", "MW"),
    "PV_to_BESS_MW":           ("D", "PV routed to BESS charging", "MW"),
    "PV_curtailed_MW":         ("D", "PV curtailed (excess not used)", "MW"),
    "BESS_charge_MW":          ("E", "BESS charge from grid/PSP", "MW"),
    "BESS_total_charge_MW":    ("E", "Total BESS charge (grid + PV)", "MW"),
    "BESS_discharge_MW":       ("E", "BESS discharge to grid", "MW"),
    "BESS_SOC_MWh":            ("E", "BESS state of charge (absolute)", "MWh"),
    "BESS_SOC_pct":            ("E", "BESS state of charge (% of 2.0 MWh capacity)", "%"),
    "Reservoir_upper_hm3":     ("F", "Upper reservoir volume end of hour", "hm³"),
    "Reservoir_lower_hm3":     ("F", "Lower reservoir volume end of hour", "hm³"),
    "Reservoir_upper_pct":     ("F", "Upper fill % (0%=830 hm³ floor, 100%=3150 hm³ usable ceiling)", "%"),
    "Reservoir_lower_pct":     ("F", "Lower fill % (0%=5 hm³ floor, 100%=54 hm³ capacity)", "%"),
    "Head_net_m":              ("F", "Net hydraulic head from linear head-vol model: 54.7 + 7.89e-9×(v_up_m³ − 830e6) m. Range 54.7–73.0 m", "m"),
    "Spill_m3h":               ("F", "Spillage from upper reservoir", "m³/h"),
    "dReservoir_upper_hm3":    ("F", "Actual Δ upper volume this hour", "hm³"),
    "dReservoir_theoretical_hm3": ("F", "Theoretical Δ volume from water balance", "hm³"),
    "Mass_balance_error_hm3":  ("F", "|Actual − Theoretical| Δ volume", "hm³"),
    "Eta_turbine_pw":          ("G", "Power-weighted turbine efficiency ηt", "—"),
    "Eta_pump_pw":             ("G", "Power-weighted pump efficiency ηp", "—"),
    "CF_turbine":              ("G", "Turbine capacity factor: PSP_gen / 518.4 MW (4×129.6 PSP only, excludes PV+BESS)", "—"),
    "CF_pump":                 ("G", "Pump capacity factor: PSP_pump / 446.4 MW (4×111.6 PSP only, excludes BESS)", "—"),
    "IDA1_price_EUR_MWh":      ("H", "IDA1 auction clearing price", "EUR/MWh"),
    "IDA1_spread_EUR_MWh":     ("H", "IDA1 price − DA price (profitability signal)", "EUR/MWh"),
    "IDA1_delta_MW":           ("H", "IDA1 position change vs DA", "MW"),
    "IDA2_price_EUR_MWh":      ("H", "IDA2 auction clearing price (H1-H2 frozen — no IDA2 trade)", "EUR/MWh"),
    "IDA2_spread_EUR_MWh":     ("H", "IDA2 price − DA price (0 for H1-H2 where IDA2 did not trade)", "EUR/MWh"),
    "IDA2_delta_MW":           ("H", "IDA2 position change vs IDA1 (0 for H1-H2)", "MW"),
    "IDA3_price_EUR_MWh":      ("H", "IDA3 auction clearing price (H1-H11 frozen — no IDA3 trade)", "EUR/MWh"),
    "IDA3_spread_EUR_MWh":     ("H", "IDA3 price − DA price (0 for H1-H11 where IDA3 did not trade)", "EUR/MWh"),
    "IDA3_delta_MW":           ("H", "IDA3 position change vs IDA2 (0 for H1-H11)", "MW"),
    "XBID_delta_MW":           ("H", "XBID continuous trade delta vs IDA3", "MW"),
    "IDA_cumulative_delta_MW": ("H", "Total IDA+XBID change vs DA position", "MW"),
    "aFRR_up_MW":              ("I", "aFRR upward capacity offered (ISP sum)", "MW"),
    "aFRR_dn_MW":              ("I", "aFRR downward capacity offered", "MW"),
    "aFRR_capUp_EUR_MW":       ("I", "aFRR upward capacity price", "EUR/MW"),
    "aFRR_capDn_EUR_MW":       ("I", "aFRR downward capacity price", "EUR/MW"),
    "mFRR_up_MW":              ("J", "mFRR upward capacity offered", "MW"),
    "mFRR_dn_MW":              ("J", "mFRR downward capacity offered", "MW"),
    "mFRR_capUp_EUR_MW":       ("J", "mFRR upward capacity price", "EUR/MW"),
    "mFRR_capDn_EUR_MW":       ("J", "mFRR downward capacity price", "EUR/MW"),
    "Gen_headroom_MW":         ("K", "PR-11: p_gen_cap(524.4) − committed_net − aFRR_up − mFRR_up. ≥0 always.", "MW"),
    "Pump_headroom_MW":        ("K", "PR-11: committed_net + p_pump_cap(447.4) − aFRR_dn − mFRR_dn. ≥0 always.", "MW"),
    "Energy_balance_check_MW": ("L", "Plant_net_DA − (PSP_gen−PSP_pump+PV_used+BESS_dis−BESS_grid_chg). Must be 0; pv_to_bess excluded (internal PV→BESS bus, not grid-crossing)", "MW"),
    "Rev_DA_EUR":              ("L", "DA energy revenue this hour", "EUR"),
    "Rev_IDA_EUR":             ("L", "IDA incremental revenue vs DA", "EUR"),
    "Rev_aFRR_cap_up_EUR":     ("L", "aFRR up capacity revenue", "EUR"),
    "Rev_aFRR_cap_dn_EUR":     ("L", "aFRR dn capacity revenue", "EUR"),
    "Rev_aFRR_cap_EUR":        ("L", "aFRR total capacity revenue", "EUR"),
    "Rev_aFRR_act_EUR":        ("L", "aFRR activation energy revenue", "EUR"),
    "Rev_mFRR_cap_up_EUR":     ("L", "mFRR up capacity revenue", "EUR"),
    "Rev_mFRR_cap_dn_EUR":     ("L", "mFRR dn capacity revenue", "EUR"),
    "Rev_mFRR_cap_EUR":        ("L", "mFRR total capacity revenue", "EUR"),
    "Rev_mFRR_act_EUR":        ("L", "mFRR activation energy revenue", "EUR"),
    "Rev_imbalance_EUR":       ("L", "Imbalance settlement (dual pricing)", "EUR"),
    "Rev_hour_total_EUR":      ("L", "Total revenue this hour", "EUR"),
    "Cum_Rev_EUR":             ("L", "Cumulative revenue to this hour", "EUR"),
    "Cum_Net_MWh":             ("L", "Cumulative final net MWh to this hour", "MWh"),
}


def _repo_root() -> str:
    # daily_excel_reports/ → phase_5d_analytics_and_reporting/ → repo root (2 levels)
    here = os.path.dirname(os.path.abspath(__file__))
    return os.path.abspath(os.path.join(here, os.pardir, os.pardir))


def _autofit(ws, min_w=8, max_w=30):
    """Widen columns to roughly fit content."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_w, max(min_w, max_len + 2))


def _title_row(ws, title: str, n_cols: int):
    ws.append([title] + [""] * (n_cols - 1))
    ws.merge_cells(start_row=ws.max_row, start_column=1,
                   end_row=ws.max_row, end_column=n_cols)
    cell = ws.cell(ws.max_row, 1)
    cell.font   = _TITLE_FONT
    cell.fill   = _TITLE_FILL
    cell.alignment = _CENTER


# ── Sheet 1: Dispatch_Hourly ──────────────────────────────────────────────────
def _write_dispatch(wb: Workbook, delivery_date: str, df: pd.DataFrame):
    ws = wb.create_sheet("Dispatch_Hourly")
    cols = list(df.columns)

    _title_row(ws, f"Alqueva Dispatch — {delivery_date}", len(cols))

    # Group-band header row (row 2)
    group_row = []
    for c in cols:
        g = _COL_META.get(c, ("?",))[0]
        group_row.append(f"Group {g}")
    ws.append(group_row)
    for idx, g in enumerate(group_row, 1):
        cell = ws.cell(ws.max_row, idx)
        g_key = g.replace("Group ", "")
        cell.fill = PatternFill("solid", fgColor=_GROUP_FILLS.get(g_key, "FFEEEEEE"))
        cell.font = Font(bold=True)
        cell.alignment = _CENTER

    # Column header row (row 3)
    ws.append(cols)
    for idx, c in enumerate(cols, 1):
        cell = ws.cell(ws.max_row, idx)
        cell.font  = _HEADER_FONT
        cell.fill  = _HEADER_FILL
        cell.alignment = _CENTER

    # Data rows
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=1):
        ws.append(row)
        row_n = ws.max_row
        fill = _ALT_FILL if r_idx % 2 == 0 else None
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row_n, c_idx)
            if fill:
                cell.fill = fill
            # Flag energy/mass balance errors in red
            col_name = cols[c_idx - 1]
            if col_name == "Energy_balance_check_MW" and abs(val or 0) > 0.1:
                cell.fill = _BAD_FILL
            if col_name == "Mass_balance_error_hm3" and abs(val or 0) > 1e-4:
                cell.fill = _BAD_FILL
            if col_name in ("Gen_headroom_MW", "Pump_headroom_MW") and (val or 0) < -0.01:
                cell.fill = _BAD_FILL

    ws.freeze_panes = "A4"
    _autofit(ws, min_w=6, max_w=20)


# ── Sheet 2: ISP_Activation ───────────────────────────────────────────────────
def _write_isp_activation(wb: Workbook, delivery_date: str):
    ws = wb.create_sheet("ISP_Activation")
    act_a = ActivationStore().load(delivery_date, "aFRR")
    act_m = ActivationStore().load(delivery_date, "mFRR")

    headers = ["ISP", "Hour", "Market", "Up_MW", "Dn_MW", "Up_price_EUR_MWh",
               "Dn_price_EUR_MWh", "Duration_h", "Revenue_EUR"]
    _title_row(ws, f"ISP Activation Detail — {delivery_date}", len(headers))
    ws.append(headers)
    for c in ws[ws.max_row]:
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = _CENTER

    # Ramp-corrected ISP durations: eff_isp_h = (isp_min - fat_min/2) / 60
    # aFRR FAT=5 min:   (15-2.5)/60 = 0.208333h  (face 0.25h overstates by 20%)
    # mFRR FAT=12.5 min:(15-6.25)/60= 0.145833h  (face 0.25h overstates by 71%)
    _EFF_H_AFRR = round((15 - 5.0 / 2) / 60, 6)    # 0.208333h
    _EFF_H_MFRR = round((15 - 12.5 / 2) / 60, 6)   # 0.145833h

    isp_n = 0
    for row in sorted(act_a, key=lambda r: (r.get("isp", 0), r.get("hour", 0))):
        isp_n += 1
        dur = row.get("eff_isp_h", _EFF_H_AFRR)   # ramp-corrected aFRR default
        up  = row.get("up_mw", 0.0);  up_p = row.get("up_price_eur_mwh", 0.0)
        dn  = row.get("dn_mw", 0.0);  dn_p = row.get("dn_price_eur_mwh", 0.0)
        rev = (up * up_p + dn * dn_p) * dur
        ws.append([row.get("isp", isp_n), row.get("hour", ""),
                   "aFRR", up, dn, up_p, dn_p, dur, round(rev, 2)])
    for row in sorted(act_m, key=lambda r: (r.get("isp", 0), r.get("hour", 0))):
        isp_n += 1
        dur = row.get("eff_isp_h", _EFF_H_MFRR)   # ramp-corrected mFRR default
        up  = row.get("up_mw", 0.0);  up_p = row.get("up_price_eur_mwh", 0.0)
        dn  = row.get("dn_mw", 0.0);  dn_p = row.get("dn_price_eur_mwh", 0.0)
        rev = (up * up_p + dn * dn_p) * dur
        ws.append([row.get("isp", isp_n), row.get("hour", ""),
                   "mFRR", up, dn, up_p, dn_p, dur, round(rev, 2)])

    ws.freeze_panes = "A3"
    _autofit(ws)


# ── Sheet 3: Gate_Decisions ───────────────────────────────────────────────────
def _write_gate_decisions(wb: Workbook, delivery_date: str):
    ws = wb.create_sheet("Gate_Decisions")
    # Gate close times — CET, confirmed against market.yaml gate_close fields.
    gates = [
        ("DA",   "Day-Ahead auction (OMIE)",     "D-1 12:00 CET"),
        ("IDA1", "Intraday auction 1 (SIDC)",    "D-1 15:00 CET"),
        ("IDA2", "Intraday auction 2 (SIDC)",    "D-1 22:00 CET"),
        ("IDA3", "Intraday auction 3 (SIDC)",    "D   10:00 CET"),
        ("XBID", "Continuous intraday (XBID)",   "D   H-1 rolling"),
    ]
    headers = ["Gate", "Market description", "Gate close", "Hours active",
               "Net position MWh", "VWAP EUR/MWh", "Net revenue EUR"]
    _title_row(ws, f"Gate Decision Summary — {delivery_date}", len(headers))
    ws.append(headers)
    for c in ws[ws.max_row]:
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = _CENTER

    pstore = PositionStore()
    for gate, desc, close in gates:
        pos = pstore.load_position(delivery_date, gate)
        if not pos:
            ws.append([gate, desc, close, 0, 0.0, 0.0, 0.0])
            continue
        hours_active = sum(1 for h, v in pos.items() if abs(v.get("volume_mwh", 0)) > 0.01)
        total_mw = sum(v.get("volume_mwh", 0) for v in pos.values())
        # Volume-weighted average price (VWAP) — simple average understates if volumes vary by hour
        vwap_num = sum(abs(v.get("volume_mwh", 0)) * v.get("price_eur_mwh", 0) for v in pos.values())
        vwap_den = sum(abs(v.get("volume_mwh", 0)) for v in pos.values())
        avg_p  = vwap_num / vwap_den if vwap_den > 1e-6 else 0.0
        revenue = sum(v.get("volume_mwh", 0) * v.get("price_eur_mwh", 0)
                      for v in pos.values())
        ws.append([gate, desc, close, hours_active,
                   round(total_mw, 2), round(avg_p, 2), round(revenue, 2)])

    ws.freeze_panes = "A3"
    _autofit(ws)


# ── Sheet 4: Summary_KPIs ─────────────────────────────────────────────────────
def _write_summary_kpis(wb: Workbook, delivery_date: str,
                         kpi_rows: List[Tuple[str, str, object, str]]):
    ws = wb.create_sheet("Summary_KPIs")
    headers = ["Section", "Metric", "Value", "Unit"]
    _title_row(ws, f"KPI Summary — {delivery_date}", len(headers))
    ws.append(headers)
    for c in ws[ws.max_row]:
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = _CENTER

    for sec, metric, value, unit in kpi_rows:
        ws.append([sec, metric, value, unit])
        r = ws.max_row
        if not metric:   # blank separator
            continue
        if sec and not metric.startswith(" "):
            ws.cell(r, 1).font = _SECTION_FONT
        # colour YES/NO cells
        if value in ("YES",):
            ws.cell(r, 3).fill = _OK_FILL
        elif value in ("NO !!!",):
            ws.cell(r, 3).fill = _BAD_FILL

    ws.freeze_panes = "A3"
    _autofit(ws, min_w=10, max_w=45)
    ws.column_dimensions["C"].width = 18


# ── Sheet 5: Glossary ─────────────────────────────────────────────────────────
def _write_glossary(wb: Workbook):
    ws = wb.create_sheet("Glossary")
    headers = ["Column", "Group", "Description", "Unit"]
    _title_row(ws, "Column Glossary — Dispatch_Hourly sheet", len(headers))
    ws.append(headers)
    for c in ws[ws.max_row]:
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = _CENTER

    prev_group = None
    for col, (group, desc, unit) in _COL_META.items():
        if prev_group and group != prev_group:
            ws.append(["", "", "", ""])   # blank between groups
        ws.append([col, f"Group {group}", desc, unit])
        prev_group = group

    _autofit(ws, min_w=10, max_w=60)


# ── Public entry point ────────────────────────────────────────────────────────
def export_daily_report(delivery_date: str) -> str:
    """Build and write the 5-sheet Excel report. Returns the output file path."""
    df = build_dispatch_hourly(delivery_date)
    kpi_rows = build_summary_kpis(delivery_date, df)

    wb = Workbook()
    wb.remove(wb.active)   # remove default empty sheet

    _write_dispatch(wb, delivery_date, df)
    _write_isp_activation(wb, delivery_date)
    _write_gate_decisions(wb, delivery_date)
    _write_summary_kpis(wb, delivery_date, kpi_rows)
    _write_glossary(wb)

    out_dir = os.path.join(_repo_root(), "runtime", "reports")
    os.makedirs(out_dir, exist_ok=True)
    path = os.path.join(out_dir, f"daily_report_{delivery_date}.xlsx")
    wb.save(path)
    return path


# Backwards-compat shim: old callers passed pnl + kpis objects
def export_daily_report_legacy(delivery_date: str, pnl=None, kpis=None) -> str:
    return export_daily_report(delivery_date)
